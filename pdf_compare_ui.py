"""
Streamlit UI for PDF Structure Comparison
Provides analyst-friendly interface for comparing requirement documents
"""

import streamlit as st
import streamlit.components.v1 as components
from pdf_compare import (
    PDFStructureComparator,
    PDFComparisonAnalyzer,
    Section,
    SectionMatch
)
import html
from datetime import datetime
import json
import pandas as pd
from typing import List
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import io


def create_pdf_comparison_ui():
    """Main UI for PDF comparison"""

    # Custom CSS for PDF comparison
    st.markdown("""
    <style>
        /* Section cards */
        .section-card {
            background: #2d2d30;
            border: 1px solid #464647;
            border-radius: 6px;
            padding: 15px;
            margin: 10px 0;
            font-family: 'Segoe UI', sans-serif;
        }

        .section-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 10px;
        }

        .section-title {
            font-size: 16px;
            font-weight: bold;
            color: #cccccc;
        }

        .section-badge {
            padding: 4px 10px;
            border-radius: 4px;
            font-size: 11px;
            font-weight: bold;
            text-transform: uppercase;
        }

        .badge-unchanged { background: #858585; color: white; }
        .badge-modified { background: #d29922; color: white; }
        .badge-added { background: #57ab5a; color: white; }
        .badge-removed { background: #e5534b; color: white; }
        .badge-reordered { background: #007ACC; color: white; }

        .section-meta {
            font-size: 12px;
            color: #858585;
            margin-bottom: 10px;
        }

        .content-diff {
            background: #1e1e1e;
            border-radius: 4px;
            padding: 12px;
            margin-top: 10px;
            font-family: 'Consolas', monospace;
            font-size: 13px;
        }

        .diff-added {
            background: rgba(87, 171, 90, 0.2);
            color: #57ab5a;
            padding: 2px 4px;
            border-radius: 2px;
        }

        .diff-removed {
            background: rgba(229, 83, 75, 0.2);
            color: #f85149;
            padding: 2px 4px;
            border-radius: 2px;
            text-decoration: line-through;
        }

        .diff-modified {
            background: rgba(210, 153, 34, 0.2);
            color: #d29922;
            padding: 2px 4px;
            border-radius: 2px;
        }

        .side-by-side {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 15px;
            margin: 15px 0;
        }

        .side-panel {
            background: #1e1e1e;
            border: 1px solid #464647;
            border-radius: 6px;
            padding: 15px;
        }

        .panel-header {
            background: #2d2d30;
            color: #969696;
            padding: 8px 12px;
            margin: -15px -15px 15px -15px;
            border-radius: 6px 6px 0 0;
            font-weight: bold;
            font-size: 13px;
        }

        .content-text {
            color: #d4d4d4;
            line-height: 1.6;
            white-space: pre-wrap;
            word-wrap: break-word;
        }

        .stats-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 15px;
            margin: 20px 0;
        }

        .stat-box {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            border-radius: 8px;
            text-align: center;
            color: white;
        }

        .stat-box.modified { background: linear-gradient(135deg, #d29922 0%, #b87c1b 100%); }
        .stat-box.added { background: linear-gradient(135deg, #57ab5a 0%, #40916c 100%); }
        .stat-box.removed { background: linear-gradient(135deg, #e5534b 0%, #c73e36 100%); }
        .stat-box.reordered { background: linear-gradient(135deg, #007ACC 0%, #005a9e 100%); }

        .stat-number {
            font-size: 36px;
            font-weight: bold;
            margin-bottom: 5px;
        }

        .stat-label {
            font-size: 13px;
            opacity: 0.9;
        }

        .match-score {
            display: inline-block;
            padding: 2px 8px;
            background: #007ACC;
            color: white;
            border-radius: 3px;
            font-size: 11px;
            font-weight: bold;
        }

        .alert-critical {
            background: rgba(229, 83, 75, 0.1);
            border-left: 4px solid #e5534b;
            padding: 15px;
            margin: 10px 0;
            border-radius: 4px;
        }

        .alert-warning {
            background: rgba(210, 153, 34, 0.1);
            border-left: 4px solid #d29922;
            padding: 15px;
            margin: 10px 0;
            border-radius: 4px;
        }

        .collapsible-section {
            cursor: pointer;
            user-select: none;
        }

        .hierarchy-indicator {
            display: inline-block;
            width: 20px;
            text-align: center;
            color: #007ACC;
            font-weight: bold;
        }
    </style>
    """, unsafe_allow_html=True)

    st.title("📄 PDF Structure Comparison")
    st.markdown("Intelligent comparison of document structures - Perfect for requirement document analysis")

    # Sidebar
    with st.sidebar:
        st.header("📁 Upload Documents")

        original_pdf = st.file_uploader(
            "Original Template",
            type=['pdf'],
            help="Upload the original requirement/guideline template",
            key="pdf_original"
        )

        modified_pdf = st.file_uploader(
            "Supplier Modified Version",
            type=['pdf'],
            help="Upload the modified version from supplier",
            key="pdf_modified"
        )

        st.divider()

        st.header("⚙️ Comparison Settings")

        view_mode = st.radio(
            "View Mode",
            ["Structured Overview", "Side-by-Side", "Change List Only", "Summary Dashboard"],
            help="Choose how to visualize changes"
        )

        show_unchanged = st.checkbox("Show unchanged sections", value=False,
                                    help="Include sections with no changes")

        show_content = st.checkbox("Show full content", value=False,
                                  help="Display full section content (may be long)")

        st.divider()

        st.header("🔍 Critical Keywords")
        critical_keywords_input = st.text_area(
            "Keywords to highlight",
            value="security\nrequirement\ncompliance\nmandatory\nshall\nmust",
            help="One keyword per line. Sections with these keywords will be flagged."
        )
        critical_keywords = [kw.strip() for kw in critical_keywords_input.split('\n') if kw.strip()]

        st.divider()

        if st.button("🔍 Compare Documents", type="primary",
                    disabled=not (original_pdf and modified_pdf)):
            if original_pdf and modified_pdf:
                with st.spinner("Analyzing document structures..."):
                    perform_comparison(original_pdf, modified_pdf, critical_keywords)

    # Main content area
    if 'pdf_comparison_done' in st.session_state and st.session_state['pdf_comparison_done']:
        display_comparison_results(view_mode, show_unchanged, show_content, critical_keywords)


def perform_comparison(original_pdf, modified_pdf, critical_keywords):
    """Perform the PDF comparison"""
    try:
        # Save uploaded files temporarily
        import tempfile
        import os

        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_orig:
            tmp_orig.write(original_pdf.read())
            orig_path = tmp_orig.name

        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_mod:
            tmp_mod.write(modified_pdf.read())
            mod_path = tmp_mod.name

        # Perform comparison
        comparator = PDFStructureComparator(orig_path, mod_path)
        matches = comparator.compare()

        # Find critical changes
        critical_changes = PDFComparisonAnalyzer.find_critical_changes(matches, critical_keywords)

        # Store in session state
        st.session_state['pdf_comparator'] = comparator
        st.session_state['pdf_matches'] = matches
        st.session_state['pdf_critical_changes'] = critical_changes
        st.session_state['pdf_comparison_done'] = True

        # Clean up temp files
        os.unlink(orig_path)
        os.unlink(mod_path)

        st.success(f"✅ Analysis complete! Found {len(matches)} sections.")

    except Exception as e:
        st.error(f"❌ Error during comparison: {str(e)}")
        import traceback
        st.code(traceback.format_exc())


def display_comparison_results(view_mode, show_unchanged, show_content, critical_keywords):
    """Display comparison results based on selected view mode"""

    comparator = st.session_state['pdf_comparator']
    matches = st.session_state['pdf_matches']
    critical_changes = st.session_state['pdf_critical_changes']

    # Summary Statistics
    st.markdown("### 📊 Comparison Summary")

    summary = comparator.summary

    stats_html = f"""
    <div class="stats-grid">
        <div class="stat-box">
            <div class="stat-number">{summary['total_sections_original']}</div>
            <div class="stat-label">Original Sections</div>
        </div>
        <div class="stat-box modified">
            <div class="stat-number">{summary['modified']}</div>
            <div class="stat-label">Modified</div>
        </div>
        <div class="stat-box added">
            <div class="stat-number">{summary['added']}</div>
            <div class="stat-label">Added</div>
        </div>
        <div class="stat-box removed">
            <div class="stat-number">{summary['removed']}</div>
            <div class="stat-label">Removed</div>
        </div>
        <div class="stat-box reordered">
            <div class="stat-number">{summary['reordered']}</div>
            <div class="stat-label">Reordered</div>
        </div>
        <div class="stat-box">
            <div class="stat-number">{summary['unchanged']}</div>
            <div class="stat-label">Unchanged</div>
        </div>
    </div>
    """
    st.markdown(stats_html, unsafe_allow_html=True)

    # Critical changes alert
    if critical_changes:
        st.markdown(f"""
        <div class="alert-critical">
            <strong>⚠️ CRITICAL CHANGES DETECTED</strong><br>
            {len(critical_changes)} sections containing critical keywords have been modified or removed.
            <strong>Review carefully!</strong>
        </div>
        """, unsafe_allow_html=True)

    # Display based on view mode
    if view_mode == "Structured Overview":
        display_structured_overview(matches, show_unchanged, show_content, critical_keywords)

    elif view_mode == "Side-by-Side":
        display_side_by_side(matches, show_unchanged, critical_keywords)

    elif view_mode == "Change List Only":
        display_change_list(matches, critical_keywords)

    elif view_mode == "Summary Dashboard":
        display_summary_dashboard(comparator, matches, critical_changes)

    # Export options
    display_export_options(comparator, matches, critical_keywords)


def display_structured_overview(matches, show_unchanged, show_content, critical_keywords):
    """Display hierarchical structured view"""
    st.markdown("### 📑 Structured Document Comparison")

    # Filter matches based on settings
    filtered_matches = [m for m in matches if show_unchanged or m.change_type != 'unchanged']

    if not filtered_matches:
        st.info("No changes detected or all changes filtered out.")
        return

    for match in filtered_matches:
        render_section_card(match, show_content, critical_keywords)


def render_section_card(match: SectionMatch, show_content: bool, critical_keywords: List[str]):
    """Render a single section comparison card"""

    # Determine which section to use for title
    section = match.modified_section or match.original_section
    if not section:
        return

    # Check if critical
    is_critical = any(
        kw.lower() in (section.title + ' ' + section.content).lower()
        for kw in critical_keywords
    )

    # Level indicator
    level_indicator = "├─" * (section.level - 1) + ("└─" if section.level > 1 else "")

    # Build card HTML
    card_html = f"""
    <div class="section-card" style="{'border-left: 4px solid #e5534b;' if is_critical else ''}">
        <div class="section-header">
            <div class="section-title">
                <span class="hierarchy-indicator">{level_indicator} Lvl {section.level}</span>
                {html.escape(section.title)}
                {' 🔴 CRITICAL' if is_critical else ''}
            </div>
            <span class="section-badge badge-{match.change_type}">{match.change_type}</span>
        </div>
    """

    # Add metadata
    if match.original_section and match.modified_section:
        card_html += f"""
        <div class="section-meta">
            📄 Original: Page {match.original_section.page_number} →
            Modified: Page {match.modified_section.page_number}
            <span class="match-score">Match: {match.match_score:.0f}%</span>
        </div>
        """
    elif match.original_section:
        card_html += f"""
        <div class="section-meta">📄 Original: Page {match.original_section.page_number}</div>
        """
    elif match.modified_section:
        card_html += f"""
        <div class="section-meta">📄 Modified: Page {match.modified_section.page_number}</div>
        """

    card_html += "</div>"

    st.markdown(card_html, unsafe_allow_html=True)

    # Show content if requested
    if show_content and (match.original_section or match.modified_section):
        with st.expander("📖 View Content", expanded=False):
            if match.change_type in ['modified', 'reordered'] and match.content_changes:
                st.markdown("**Content Changes:**")
                for change_type, old_text, new_text in match.content_changes[:5]:  # Limit to 5
                    if change_type == 'modified':
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown(f"<div class='diff-removed'>{html.escape(old_text[:200])}</div>",
                                      unsafe_allow_html=True)
                        with col2:
                            st.markdown(f"<div class='diff-added'>{html.escape(new_text[:200])}</div>",
                                      unsafe_allow_html=True)
                    elif change_type == 'deleted':
                        st.markdown(f"<div class='diff-removed'>{html.escape(old_text[:200])}</div>",
                                  unsafe_allow_html=True)
                    elif change_type == 'added':
                        st.markdown(f"<div class='diff-added'>{html.escape(new_text[:200])}</div>",
                                  unsafe_allow_html=True)

                if len(match.content_changes) > 5:
                    st.info(f"... and {len(match.content_changes) - 5} more changes")
            else:
                # Show original or modified content
                content = (match.modified_section or match.original_section).content
                st.text_area("Content", content[:1000], height=200, disabled=True)


def display_side_by_side(matches, show_unchanged, critical_keywords):
    """Display side-by-side comparison"""
    st.markdown("### 🔄 Side-by-Side Comparison")

    filtered_matches = [m for m in matches if show_unchanged or m.change_type != 'unchanged']

    for match in filtered_matches:
        section = match.modified_section or match.original_section
        if not section:
            continue

        is_critical = any(
            kw.lower() in (section.title + ' ' + section.content).lower()
            for kw in critical_keywords
        )

        st.markdown(f"""
        <div style="{'border-left: 4px solid #e5534b; padding-left: 10px;' if is_critical else ''}">
            <strong>{'🔴 CRITICAL - ' if is_critical else ''}{html.escape(section.title)}</strong>
            <span class="section-badge badge-{match.change_type}">{match.change_type}</span>
        </div>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**📁 Original**")
            if match.original_section:
                st.markdown(f"*Page {match.original_section.page_number}*")
                st.text_area(
                    "Original Content",
                    match.original_section.content[:500],
                    height=150,
                    disabled=True,
                    key=f"orig_{match.original_section.section_id}"
                )
            else:
                st.info("Section not present in original")

        with col2:
            st.markdown("**📝 Modified**")
            if match.modified_section:
                st.markdown(f"*Page {match.modified_section.page_number}*")
                st.text_area(
                    "Modified Content",
                    match.modified_section.content[:500],
                    height=150,
                    disabled=True,
                    key=f"mod_{match.modified_section.section_id}"
                )
            else:
                st.warning("Section removed in modified version")

        st.markdown("---")


def display_change_list(matches, critical_keywords):
    """Display compact change list"""
    st.markdown("### 📋 Change List")

    changes_only = [m for m in matches if m.change_type != 'unchanged']

    if not changes_only:
        st.success("✅ No changes detected!")
        return

    for idx, match in enumerate(changes_only, 1):
        section = match.modified_section or match.original_section

        is_critical = any(
            kw.lower() in (section.title + ' ' + section.content).lower()
            for kw in critical_keywords
        )

        change_html = f"""
        <div class="section-card" style="{'border-left: 4px solid #e5534b;' if is_critical else ''}">
            <strong>{idx}. {'🔴 ' if is_critical else ''}{html.escape(section.title)}</strong>
            <span class="section-badge badge-{match.change_type}">{match.change_type}</span>
            <br>
            <small>Level {section.level} | Page {section.page_number}</small>
        </div>
        """
        st.markdown(change_html, unsafe_allow_html=True)


def display_summary_dashboard(comparator, matches, critical_changes):
    """Display analytical summary dashboard"""
    st.markdown("### 📊 Analysis Dashboard")

    # Create summary DataFrame
    summary_data = []

    change_type_counts = {
        'unchanged': 0,
        'modified': 0,
        'added': 0,
        'removed': 0,
        'reordered': 0
    }

    level_stats = {}

    for match in matches:
        change_type_counts[match.change_type] += 1

        section = match.modified_section or match.original_section
        if section:
            level = section.level
            if level not in level_stats:
                level_stats[level] = {'unchanged': 0, 'modified': 0, 'added': 0, 'removed': 0, 'reordered': 0}
            level_stats[level][match.change_type] += 1

    # Level-wise breakdown
    st.markdown("#### 📊 Changes by Hierarchy Level")

    level_df_data = []
    for level in sorted(level_stats.keys()):
        level_name = {1: 'Chapter', 2: 'Section', 3: 'Subsection', 4: 'Sub-subsection'}.get(level, f'Level {level}')
        stats = level_stats[level]
        level_df_data.append({
            'Level': level_name,
            'Modified': stats['modified'],
            'Added': stats['added'],
            'Removed': stats['removed'],
            'Reordered': stats['reordered'],
            'Unchanged': stats['unchanged'],
            'Total': sum(stats.values())
        })

    if level_df_data:
        level_df = pd.DataFrame(level_df_data)
        st.dataframe(level_df, use_container_width=True, hide_index=True)

    # Critical changes
    if critical_changes:
        st.markdown("#### ⚠️ Critical Changes Requiring Review")

        critical_df_data = []
        for match in critical_changes:
            section = match.modified_section or match.original_section
            critical_df_data.append({
                'Section': section.title[:50],
                'Level': section.level,
                'Page': section.page_number,
                'Change Type': match.change_type,
                'Content Changes': len(match.content_changes) if match.content_changes else 0
            })

        critical_df = pd.DataFrame(critical_df_data)
        st.dataframe(critical_df, use_container_width=True, hide_index=True)


def display_export_options(comparator, matches, critical_keywords):
    """Display export options"""
    st.markdown("---")
    st.markdown("### 💾 Export Options")

    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("📄 Export Text Report", type="secondary"):
            report = PDFComparisonAnalyzer.generate_change_report(comparator)

            st.download_button(
                label="⬇️ Download Report",
                data=report,
                file_name=f"pdf_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                mime="text/plain"
            )

    with col2:
        if st.button("📊 Export Excel Summary", type="secondary"):
            excel_file = export_to_excel(comparator, matches, critical_keywords)

            st.download_button(
                label="⬇️ Download Excel",
                data=excel_file,
                file_name=f"pdf_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with col3:
        if st.button("📋 Export JSON", type="secondary"):
            json_data = json.dumps(comparator.export_to_dict(), indent=2, default=str)

            st.download_button(
                label="⬇️ Download JSON",
                data=json_data,
                file_name=f"pdf_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )


def export_to_excel(comparator, matches, critical_keywords):
    """Export comparison to Excel with formatting"""
    wb = openpyxl.Workbook()

    # Summary sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    summary_data = [
        ["PDF Structure Comparison Report", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Metric", "Count"],
        ["Original Sections", comparator.summary['total_sections_original']],
        ["Modified Sections", comparator.summary['total_sections_modified']],
        ["Unchanged", comparator.summary['unchanged']],
        ["Modified", comparator.summary['modified']],
        ["Added", comparator.summary['added']],
        ["Removed", comparator.summary['removed']],
        ["Reordered", comparator.summary['reordered']],
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 4:
                cell.font = Font(bold=True)

    # Detailed changes sheet
    changes_sheet = wb.create_sheet("Detailed Changes")
    changes_sheet.append(["Section Title", "Level", "Page (Orig)", "Page (Mod)", "Change Type", "Match Score", "Content Changes"])

    # Color fills
    fills = {
        'modified': PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
        'added': PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        'removed': PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        'reordered': PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid"),
    }

    for match in matches:
        if match.change_type == 'unchanged':
            continue

        section = match.modified_section or match.original_section
        orig_page = match.original_section.page_number if match.original_section else "N/A"
        mod_page = match.modified_section.page_number if match.modified_section else "N/A"

        row = [
            section.title,
            section.level,
            orig_page,
            mod_page,
            match.change_type,
            f"{match.match_score:.1f}%" if match.match_score > 0 else "N/A",
            len(match.content_changes) if match.content_changes else 0
        ]

        changes_sheet.append(row)

        # Apply formatting
        if match.change_type in fills:
            for cell in changes_sheet[changes_sheet.max_row]:
                cell.fill = fills[match.change_type]

    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()


if __name__ == "__main__":
    st.set_page_config(
        page_title="PDF Structure Comparison",
        page_icon="📄",
        layout="wide"
    )

    create_pdf_comparison_ui()
