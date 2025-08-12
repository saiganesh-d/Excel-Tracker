import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
from datetime import datetime
import json
import io
import base64
from pathlib import Path
import difflib
import html
import streamlit.components.v1 as components

# Page configuration
st.set_page_config(
    page_title="Excel Diff Visualizer",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for VS Code-style diff visualization
st.markdown("""
<style>
    .diff-container {
        font-family: 'Monaco', 'Menlo', 'Ubuntu Mono', monospace;
        font-size: 13px;
        background: #1e1e1e;
        border-radius: 6px;
        padding: 10px;
        margin: 10px 0;
        overflow: hidden;
    }
    
    .diff-header {
        background: #2d2d30;
        color: #cccccc;
        padding: 8px 12px;
        border-radius: 4px 4px 0 0;
        font-weight: bold;
        border-bottom: 1px solid #464647;
    }
    
    .diff-grid {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 2px;
        background: #2d2d30;
        padding: 1px;
    }
    
    .diff-side {
        background: #1e1e1e;
        overflow: auto;
        max-height: 500px;
        max-width: 100%;
    }
    
    .diff-side-header {
        background: #2d2d30;
        color: #969696;
        padding: 6px 12px;
        font-size: 12px;
        border-bottom: 1px solid #464647;
        position: sticky;
        top: 0;
        z-index: 10;
    }
    
    .diff-content {
        overflow-x: auto;
        overflow-y: auto;
        max-height: 450px;
        width: 100%;
    }
    
    .diff-table {
        display: table;
        width: max-content;
        min-width: 100%;
        border-collapse: collapse;
    }
    
    .diff-line {
        display: table-row;
        min-height: 22px;
        white-space: nowrap;
    }
    
    .line-number {
        display: table-cell;
        background: #2d2d30;
        color: #858585;
        padding: 2px 8px;
        min-width: 40px;
        text-align: right;
        border-right: 1px solid #464647;
        user-select: none;
        position: sticky;
        left: 0;
        z-index: 5;
    }
    
    .line-content {
        display: table-cell;
        padding: 2px 0;
        white-space: nowrap;
    }
    
    .cell-data {
        display: inline-block;
        padding: 2px 8px;
        min-width: 100px;
        border-right: 1px solid #3a3a3a;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        max-width: 200px;
    }
    
    .cell-data:hover {
        max-width: none;
        background: rgba(255, 255, 255, 0.05);
        position: relative;
        z-index: 2;
    }
    
    /* Diff highlighting colors - VS Code style */
    .diff-modified {
        background: rgba(210, 153, 34, 0.15) !important;
    }
    
    .diff-modified .line-number {
        background: #3d3319 !important;
        color: #d29922;
    }
    
    .diff-unchanged {
        background: #1e1e1e;
        color: #d4d4d4;
    }
    
    /* Cell highlighting for modifications */
    .cell-empty {
        color: #6a6a6a;
        font-style: italic;
        opacity: 0.6;
    }
    
    .cell-value-old {
        background: rgba(229, 83, 75, 0.2) !important;
        color: #f85149 !important;
        padding: 1px 4px;
        border-radius: 2px;
        text-decoration: line-through;
    }
    
    .cell-value-new {
        background: rgba(87, 171, 90, 0.2) !important;
        color: #57ab5a !important;
        padding: 1px 4px;
        border-radius: 2px;
        font-weight: bold;
    }
    
    /* Column headers */
    .header-row {
        background: #2d2d30 !important;
        font-weight: bold;
        position: sticky;
        top: 0;
        z-index: 6;
        border-bottom: 2px solid #464647;
    }
    
    .header-row .cell-data {
        background: #2d2d30;
        color: #007ACC;
        font-weight: bold;
        border-right: 1px solid #464647;
    }
    
    /* Stats cards */
    .stats-container {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
        gap: 15px;
        margin: 20px 0;
    }
    
    .stat-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px;
        border-radius: 10px;
        color: white;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    
    .stat-card.modified {
        background: linear-gradient(135deg, #d29922 0%, #b87c1b 100%);
    }
    
    .stat-card.to-empty {
        background: linear-gradient(135deg, #e5534b 0%, #c73e36 100%);
    }
    
    .stat-card.from-empty {
        background: linear-gradient(135deg, #57ab5a 0%, #40916c 100%);
    }
    
    .stat-number {
        font-size: 32px;
        font-weight: bold;
        margin-bottom: 5px;
    }
    
    .stat-label {
        font-size: 14px;
        opacity: 0.9;
    }
    
    /* Sync scroll indicator */
    .sync-indicator {
        background: #007ACC;
        color: white;
        padding: 4px 8px;
        border-radius: 4px;
        font-size: 11px;
        margin-left: 10px;
    }
    
    /* Change list styling */
    .change-item {
        background: #2d2d30;
        border: 1px solid #464647;
        border-radius: 4px;
        padding: 10px;
        margin: 5px 0;
        font-family: monospace;
    }
    
    .change-location {
        color: #007ACC;
        font-weight: bold;
    }
    
    .change-arrow {
        color: #d29922;
        margin: 0 8px;
    }
    
    .value-empty {
        color: #6a6a6a;
        font-style: italic;
    }
    
    .value-old {
        color: #f85149;
        text-decoration: line-through;
    }
    
    .value-new {
        color: #57ab5a;
        font-weight: bold;
    }
    
    /* Scrollbar styling */
    .diff-content::-webkit-scrollbar {
        width: 10px;
        height: 10px;
    }
    
    .diff-content::-webkit-scrollbar-track {
        background: #2d2d30;
    }
    
    .diff-content::-webkit-scrollbar-thumb {
        background: #464647;
        border-radius: 5px;
    }
    
    .diff-content::-webkit-scrollbar-thumb:hover {
        background: #565658;
    }
    
    .diff-content::-webkit-scrollbar-corner {
        background: #2d2d30;
    }
</style>
""", unsafe_allow_html=True)

class ExcelDiffVisualizer:
    def __init__(self, original_file, modified_file):
        self.original_wb = openpyxl.load_workbook(original_file, data_only=True)
        self.modified_wb = openpyxl.load_workbook(modified_file, data_only=True)
        self.changes = {}
        self.summary = {
            'total_modifications': 0,
            'sheets_modified': [],
            'blank_to_value': 0,  # Previously "added"
            'value_to_blank': 0,  # Previously "removed"
            'value_to_value': 0,  # Previously "modified"
        }
    
    def get_sheet_as_dataframe(self, sheet):
        """Convert sheet to DataFrame for easier comparison"""
        data = []
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Store headers separately (assuming first row contains headers)
        headers = {}
        if max_row >= 1:
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=1, column=col)
                if cell.value:
                    headers[col] = str(cell.value)
                else:
                    headers[col] = get_column_letter(col)
        
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                # Store None for empty cells to distinguish from empty string
                row_data.append(cell.value if cell.value is not None else None)
            data.append(row_data)
        
        # Create column headers
        columns = [get_column_letter(i) for i in range(1, max_col + 1)]
        
        if data:
            df = pd.DataFrame(data, columns=columns[:len(data[0])])
        else:
            df = pd.DataFrame()
        
        # Store headers for later use
        self.column_headers = headers
        
        return df
    
    def compare_sheets(self):
        """Compare all sheets in the workbooks"""
        all_sheets = set(self.original_wb.sheetnames) | set(self.modified_wb.sheetnames)
        
        for sheet_name in all_sheets:
            sheet_changes = {
                'modifications': [],  # All changes are now modifications
                'sheet_added': False,
                'sheet_removed': False,
                'original_df': None,
                'modified_df': None,
                'column_headers': {}
            }
            
            # Check if sheet exists in both workbooks
            if sheet_name not in self.original_wb.sheetnames:
                sheet_changes['sheet_added'] = True
                sheet_changes['modified_df'] = self.get_sheet_as_dataframe(self.modified_wb[sheet_name])
                sheet_changes['column_headers'] = getattr(self, 'column_headers', {})
                self.summary['sheets_modified'].append(f"{sheet_name} (New Sheet)")
            elif sheet_name not in self.modified_wb.sheetnames:
                sheet_changes['sheet_removed'] = True
                sheet_changes['original_df'] = self.get_sheet_as_dataframe(self.original_wb[sheet_name])
                sheet_changes['column_headers'] = getattr(self, 'column_headers', {})
                self.summary['sheets_modified'].append(f"{sheet_name} (Sheet Removed)")
            else:
                # Get both sheets as DataFrames
                original_sheet = self.original_wb[sheet_name]
                modified_sheet = self.modified_wb[sheet_name]
                
                sheet_changes['original_df'] = self.get_sheet_as_dataframe(original_sheet)
                original_headers = dict(getattr(self, 'column_headers', {}))
                
                sheet_changes['modified_df'] = self.get_sheet_as_dataframe(modified_sheet)
                modified_headers = dict(getattr(self, 'column_headers', {}))
                
                # Use modified headers as primary, fall back to original if needed
                sheet_changes['column_headers'] = modified_headers or original_headers
                
                # Get max dimensions for both sheets
                max_row = max(original_sheet.max_row, modified_sheet.max_row, 10)  # At least 10 rows
                max_col = max(original_sheet.max_column, modified_sheet.max_column, 5)  # At least 5 columns
                
                # Compare all cells in the grid
                has_changes = False
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        orig_cell = original_sheet.cell(row=row, column=col)
                        mod_cell = modified_sheet.cell(row=row, column=col)
                        
                        orig_val = orig_cell.value
                        mod_val = mod_cell.value
                        
                        # Only track if there's an actual change
                        if orig_val != mod_val:
                            change_type = self._categorize_change(orig_val, mod_val)
                            
                            # Get column header name for better description
                            col_header = sheet_changes['column_headers'].get(col, get_column_letter(col))
                            
                            sheet_changes['modifications'].append({
                                'cell': (row, col),
                                'cell_ref': f"{get_column_letter(col)}{row}",
                                'column_name': col_header,
                                'row_number': row,
                                'old_value': orig_val,
                                'new_value': mod_val,
                                'row': row,
                                'col': col,
                                'change_type': change_type
                            })
                            
                            # Update summary
                            if change_type == 'blank_to_value':
                                self.summary['blank_to_value'] += 1
                            elif change_type == 'value_to_blank':
                                self.summary['value_to_blank'] += 1
                            else:
                                self.summary['value_to_value'] += 1
                            
                            has_changes = True
                            self.summary['total_modifications'] += 1
                
                if has_changes:
                    self.summary['sheets_modified'].append(sheet_name)
            
            self.changes[sheet_name] = sheet_changes
        
        return self.changes
    
    def _categorize_change(self, old_value, new_value):
        """Categorize the type of modification"""
        if old_value is None and new_value is not None:
            return 'blank_to_value'
        elif old_value is not None and new_value is None:
            return 'value_to_blank'
        else:
            return 'value_to_value'
    
    def _format_value(self, value):
        """Format value for display"""
        if value is None:
            return "[empty]"
        elif isinstance(value, (int, float)):
            return str(value)
        else:
            return str(value)
    
    def create_vs_code_diff_html(self, sheet_name):
        """Create VS Code style diff visualization for a sheet"""
        sheet_changes = self.changes.get(sheet_name, {})
        
        if sheet_changes.get('sheet_added'):
            return self._create_added_sheet_view(sheet_changes['modified_df'], sheet_name)
        elif sheet_changes.get('sheet_removed'):
            return self._create_removed_sheet_view(sheet_changes['original_df'], sheet_name)
        
        original_df = sheet_changes.get('original_df', pd.DataFrame())
        modified_df = sheet_changes.get('modified_df', pd.DataFrame())
        column_headers = sheet_changes.get('column_headers', {})
        
        # Get max dimensions
        max_rows = max(len(original_df) if not original_df.empty else 10,
                      len(modified_df) if not modified_df.empty else 10,
                      10)
        max_cols = max(len(original_df.columns) if not original_df.empty else 5,
                      len(modified_df.columns) if not modified_df.empty else 5,
                      5)
        
        # Build change map
        change_map = {}
        for mod in sheet_changes.get('modifications', []):
            change_map[(mod['row']-1, mod['col']-1)] = mod
        
        # Generate unique ID for this sheet's containers
        sheet_id = sheet_name.replace(' ', '_').replace('(', '').replace(')', '')
        
        html = f"""
        <div class="diff-container">
            <div class="diff-header">
                📊 Sheet: {sheet_name}
                <span style="background: #007ACC; color: white; padding: 4px 8px; border-radius: 4px; font-size: 11px; margin-left: 10px;">
                    📜 Scroll to compare
                </span>
            </div>
            <div class="diff-grid">
                <div class="diff-side">
                    <div class="diff-side-header">📁 Original</div>
                    <div class="diff-content">
                        <div class="diff-table">
        """
        
        # Add column headers for original
        html += '<div class="diff-line header-row"><span class="line-number">⬜</span><span class="line-content">'
        for col_idx in range(max_cols):
            col_num = col_idx + 1
            if col_num in column_headers:
                header = column_headers[col_num]
            elif col_idx < len(original_df.columns):
                header = original_df.columns[col_idx]
            else:
                header = get_column_letter(col_num)
            html += f'<span class="cell-data">{html.escape(str(header))}</span>'
        html += '</span></div>'
        
        # Original side - data rows
        for row_idx in range(max_rows):
            row_has_changes = any((row_idx, col) in change_map for col in range(max_cols))
            
            html += f'<div class="diff-line'
            if row_has_changes:
                html += ' diff-modified'
            html += f'"><span class="line-number">{row_idx + 1}</span><span class="line-content">'
            
            for col_idx in range(max_cols):
                cell_html = '<span class="cell-data">'
                
                if row_idx < len(original_df) and col_idx < len(original_df.columns):
                    value = original_df.iloc[row_idx, col_idx]
                    
                    if (row_idx, col_idx) in change_map:
                        change = change_map[(row_idx, col_idx)]
                        if value is None:
                            cell_html += '<span class="cell-empty">[empty]</span>'
                        else:
                            cell_html += f'<span class="cell-value-old">{html.escape(str(value))}</span>'
                    else:
                        if value is None:
                            cell_html += '<span class="cell-empty">·</span>'
                        else:
                            cell_html += html.escape(str(value))
                else:
                    cell_html += '<span class="cell-empty">·</span>'
                
                cell_html += '</span>'
                html += cell_html
            
            html += '</span></div>'
        
        html += """
                        </div>
                    </div>
                </div>
                <div class="diff-side">
                    <div class="diff-side-header">📝 Modified</div>
                    <div class="diff-content">
                        <div class="diff-table">
        """
        
        # Add column headers for modified
        html += '<div class="diff-line header-row"><span class="line-number">⬜</span><span class="line-content">'
        for col_idx in range(max_cols):
            col_num = col_idx + 1
            if col_num in column_headers:
                header = column_headers[col_num]
            elif col_idx < len(modified_df.columns):
                header = modified_df.columns[col_idx]
            else:
                header = get_column_letter(col_num)
            html += f'<span class="cell-data">{html.escape(str(header))}</span>'
        html += '</span></div>'
        
        # Modified side - data rows
        for row_idx in range(max_rows):
            row_has_changes = any((row_idx, col) in change_map for col in range(max_cols))
            
            html += f'<div class="diff-line'
            if row_has_changes:
                html += ' diff-modified'
            html += f'"><span class="line-number">{row_idx + 1}</span><span class="line-content">'
            
            for col_idx in range(max_cols):
                cell_html = '<span class="cell-data">'
                
                if row_idx < len(modified_df) and col_idx < len(modified_df.columns):
                    value = modified_df.iloc[row_idx, col_idx]
                    
                    if (row_idx, col_idx) in change_map:
                        change = change_map[(row_idx, col_idx)]
                        if value is None:
                            cell_html += '<span class="cell-empty">[empty]</span>'
                        else:
                            cell_html += f'<span class="cell-value-new">{html.escape(str(value))}</span>'
                    else:
                        if value is None:
                            cell_html += '<span class="cell-empty">·</span>'
                        else:
                            cell_html += html.escape(str(value))
                else:
                    cell_html += '<span class="cell-empty">·</span>'
                
                cell_html += '</span>'
                html += cell_html
            
            html += '</span></div>'
        
        html += """
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """
        
        return html
    
    def _create_added_sheet_view(self, df, sheet_name):
        """Create view for entirely added sheet"""
        html = f"""
        <div class="diff-container">
            <div class="diff-header" style="background: #57ab5a; color: white;">
                📄 New Sheet: {sheet_name}
            </div>
            <div style="padding: 20px; background: rgba(87, 171, 90, 0.1); color: #d4d4d4;">
                <p style="margin-bottom: 15px;">This entire sheet was added. All cells are modifications from [empty] to their current values.</p>
        """
        
        if not df.empty:
            for row_idx in range(min(len(df), 20)):  # Show first 20 rows
                html += f'<div class="diff-line diff-modified">'
                html += f'<span class="line-number">{row_idx + 1}</span>'
                html += '<span class="line-content">'
                row_data = []
                for col_idx in range(len(df.columns)):
                    value = df.iloc[row_idx, col_idx]
                    if value is not None:
                        row_data.append(f'<span class="value-new">{html.escape(str(value))}</span>')
                    else:
                        row_data.append('<span class="cell-empty">[empty]</span>')
                html += ' | '.join(row_data)
                html += '</span></div>'
            
            if len(df) > 20:
                html += f'<div style="padding: 10px; color: #858585;">... and {len(df) - 20} more rows</div>'
        
        html += "</div></div>"
        return html
    
    def create_synchronized_diff_component(self, sheet_name):
        """Create synchronized diff view using Streamlit components for true scrolling sync"""
        sheet_changes = self.changes.get(sheet_name, {})
        
        if sheet_changes.get('sheet_added') or sheet_changes.get('sheet_removed'):
            # For added/removed sheets, use the regular view
            return self.create_vs_code_diff_html(sheet_name)
        
        original_df = sheet_changes.get('original_df', pd.DataFrame())
        modified_df = sheet_changes.get('modified_df', pd.DataFrame())
        column_headers = sheet_changes.get('column_headers', {})
        
        # Get max dimensions
        max_rows = max(len(original_df) if not original_df.empty else 10,
                      len(modified_df) if not modified_df.empty else 10, 10)
        max_cols = max(len(original_df.columns) if not original_df.empty else 5,
                      len(modified_df.columns) if not modified_df.empty else 5, 5)
        
        # Build change map
        change_map = {}
        for mod in sheet_changes.get('modifications', []):
            change_map[(mod['row']-1, mod['col']-1)] = mod
        
        # Clean sheet name for IDs
        sheet_id = sheet_name.replace(' ', '_').replace('(', '').replace(')', '')
        
        # Generate HTML for original side
        original_html = self._generate_table_html(
            original_df, column_headers, change_map, max_rows, max_cols, is_original=True
        )
        
        # Generate HTML for modified side
        modified_html = self._generate_table_html(
            modified_df, column_headers, change_map, max_rows, max_cols, is_original=False
        )
        
        # Create full HTML with embedded JavaScript for synchronized scrolling
        full_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{
                    margin: 0;
                    padding: 0;
                    font-family: 'Monaco', 'Menlo', 'Ubuntu Mono', monospace;
                    background: #1e1e1e;
                    color: #d4d4d4;
                }}
                
                .diff-container {{
                    background: #1e1e1e;
                    border-radius: 6px;
                    overflow: hidden;
                    height: 100vh;
                    display: flex;
                    flex-direction: column;
                }}
                
                .diff-header {{
                    background: #2d2d30;
                    color: #cccccc;
                    padding: 10px 15px;
                    font-weight: bold;
                    border-bottom: 1px solid #464647;
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    flex-shrink: 0;
                }}
                
                .sync-badge {{
                    background: #28a745;
                    color: white;
                    padding: 4px 10px;
                    border-radius: 4px;
                    font-size: 11px;
                    animation: pulse 2s infinite;
                }}
                
                @keyframes pulse {{
                    0% {{ opacity: 1; }}
                    50% {{ opacity: 0.7; }}
                    100% {{ opacity: 1; }}
                }}
                
                .diff-grid {{
                    display: grid;
                    grid-template-columns: 1fr 1fr;
                    gap: 2px;
                    background: #2d2d30;
                    padding: 1px;
                    flex: 1;
                    overflow: hidden;
                }}
                
                .diff-side {{
                    background: #1e1e1e;
                    display: flex;
                    flex-direction: column;
                    overflow: hidden;
                }}
                
                .diff-side-header {{
                    background: #2d2d30;
                    color: #969696;
                    padding: 8px 15px;
                    font-size: 12px;
                    border-bottom: 1px solid #464647;
                    flex-shrink: 0;
                }}
                
                .diff-content {{
                    overflow: auto;
                    flex: 1;
                }}
                
                .diff-table {{
                    display: table;
                    width: max-content;
                    min-width: 100%;
                    border-collapse: collapse;
                }}
                
                .diff-line {{
                    display: table-row;
                    min-height: 22px;
                    white-space: nowrap;
                }}
                
                .diff-line:hover {{
                    background: rgba(255, 255, 255, 0.05) !important;
                }}
                
                .line-number {{
                    display: table-cell;
                    background: #2d2d30;
                    color: #858585;
                    padding: 2px 8px;
                    min-width: 40px;
                    text-align: right;
                    border-right: 1px solid #464647;
                    position: sticky;
                    left: 0;
                    z-index: 5;
                    user-select: none;
                }}
                
                .line-content {{
                    display: table-cell;
                    padding: 2px 0;
                    white-space: nowrap;
                }}
                
                .cell-data {{
                    display: inline-block;
                    padding: 2px 12px;
                    min-width: 120px;
                    border-right: 1px solid #3a3a3a;
                    white-space: nowrap;
                }}
                
                .header-row {{
                    background: #2d2d30 !important;
                    font-weight: bold;
                    position: sticky;
                    top: 0;
                    z-index: 6;
                }}
                
                .header-row .cell-data {{
                    background: #2d2d30;
                    color: #007ACC;
                    font-weight: bold;
                    border-right: 1px solid #464647;
                }}
                
                .diff-modified {{
                    background: rgba(210, 153, 34, 0.15) !important;
                }}
                
                .diff-modified .line-number {{
                    background: #3d3319 !important;
                    color: #d29922;
                }}
                
                .cell-empty {{
                    color: #6a6a6a;
                    font-style: italic;
                    opacity: 0.6;
                }}
                
                .cell-value-old {{
                    background: rgba(229, 83, 75, 0.2);
                    color: #f85149;
                    padding: 1px 4px;
                    border-radius: 2px;
                    text-decoration: line-through;
                }}
                
                .cell-value-new {{
                    background: rgba(87, 171, 90, 0.2);
                    color: #57ab5a;
                    padding: 1px 4px;
                    border-radius: 2px;
                    font-weight: bold;
                }}
                
                /* Scrollbar styling */
                ::-webkit-scrollbar {{
                    width: 10px;
                    height: 10px;
                }}
                
                ::-webkit-scrollbar-track {{
                    background: #2d2d30;
                }}
                
                ::-webkit-scrollbar-thumb {{
                    background: #464647;
                    border-radius: 5px;
                }}
                
                ::-webkit-scrollbar-thumb:hover {{
                    background: #565658;
                }}
                
                ::-webkit-scrollbar-corner {{
                    background: #2d2d30;
                }}
            </style>
        </head>
        <body>
            <div class="diff-container">
                <div class="diff-header">
                    📊 Sheet: {sheet_name}
                    <span class="sync-badge">⟷ SYNCHRONIZED SCROLLING ACTIVE</span>
                </div>
                <div class="diff-grid">
                    <div class="diff-side">
                        <div class="diff-side-header">📁 Original</div>
                        <div class="diff-content" id="original-{sheet_id}">
                            {original_html}
                        </div>
                    </div>
                    <div class="diff-side">
                        <div class="diff-side-header">📝 Modified</div>
                        <div class="diff-content" id="modified-{sheet_id}">
                            {modified_html}
                        </div>
                    </div>
                </div>
            </div>
            
            <script>
                // Synchronized scrolling implementation
                (function() {{
                    const original = document.getElementById('original-{sheet_id}');
                    const modified = document.getElementById('modified-{sheet_id}');
                    
                    if (original && modified) {{
                        let isSyncing = false;
                        
                        // Function to sync scroll positions
                        function syncScroll(source, target) {{
                            if (!isSyncing) {{
                                isSyncing = true;
                                target.scrollTop = source.scrollTop;
                                target.scrollLeft = source.scrollLeft;
                                
                                // Reset flag after a small delay
                                requestAnimationFrame(() => {{
                                    isSyncing = false;
                                }});
                            }}
                        }}
                        
                        // Add scroll listeners
                        original.addEventListener('scroll', function() {{
                            syncScroll(original, modified);
                        }});
                        
                        modified.addEventListener('scroll', function() {{
                            syncScroll(modified, original);
                        }});
                        
                        // Log that synchronization is active
                        console.log('✅ Synchronized scrolling activated for sheet: {sheet_name}');
                    }} else {{
                        console.error('❌ Could not find elements for synchronization');
                    }}
                }})();
            </script>
        </body>
        </html>
        """
        
        return full_html
    
    def _generate_table_html(self, df, column_headers, change_map, max_rows, max_cols, is_original=True):
        """Generate HTML table for one side of the diff"""
        html_output = '<div class="diff-table">'
        
        # Add column headers
        html_output += '<div class="diff-line header-row"><span class="line-number">⬜</span><span class="line-content">'
        for col_idx in range(max_cols):
            col_num = col_idx + 1
            if col_num in column_headers:
                header = column_headers[col_num]
            elif col_idx < len(df.columns) if not df.empty else 0:
                header = df.columns[col_idx]
            else:
                header = get_column_letter(col_num)
            html_output += f'<span class="cell-data">{html.escape(str(header))}</span>'
        html_output += '</span></div>'
        
        # Add data rows
        for row_idx in range(max_rows):
            row_has_changes = any((row_idx, col) in change_map for col in range(max_cols))
            
            html_output += f'<div class="diff-line'
            if row_has_changes:
                html_output += ' diff-modified'
            html_output += f'"><span class="line-number">{row_idx + 1}</span><span class="line-content">'
            
            for col_idx in range(max_cols):
                cell_html = '<span class="cell-data">'
                
                if not df.empty and row_idx < len(df) and col_idx < len(df.columns):
                    value = df.iloc[row_idx, col_idx]
                    
                    if (row_idx, col_idx) in change_map:
                        if value is None:
                            cell_html += '<span class="cell-empty">[empty]</span>'
                        else:
                            if is_original:
                                cell_html += f'<span class="cell-value-old">{html.escape(str(value))}</span>'
                            else:
                                cell_html += f'<span class="cell-value-new">{html.escape(str(value))}</span>'
                    else:
                        if value is None:
                            cell_html += '<span class="cell-empty">·</span>'
                        else:
                            cell_html += html.escape(str(value))
                else:
                    cell_html += '<span class="cell-empty">·</span>'
                
                cell_html += '</span>'
                html_output += cell_html
            
            html_output += '</span></div>'
        
        html_output += '</div>'
        return html_output
    
    def create_synchronized_diff_component(self, sheet_name):
        """Create synchronized diff view using Streamlit components for true scrolling sync"""
        sheet_changes = self.changes.get(sheet_name, {})
        
        if sheet_changes.get('sheet_added') or sheet_changes.get('sheet_removed'):
            # For added/removed sheets, use the regular view
            return self.create_vs_code_diff_html(sheet_name)
        
        original_df = sheet_changes.get('original_df', pd.DataFrame())
        modified_df = sheet_changes.get('modified_df', pd.DataFrame())
        column_headers = sheet_changes.get('column_headers', {})
        
        # Get max dimensions
        max_rows = max(len(original_df) if not original_df.empty else 10,
                      len(modified_df) if not modified_df.empty else 10, 10)
        max_cols = max(len(original_df.columns) if not original_df.empty else 5,
                      len(modified_df.columns) if not modified_df.empty else 5, 5)
        
        # Build change map
        change_map = {}
        for mod in sheet_changes.get('modifications', []):
            change_map[(mod['row']-1, mod['col']-1)] = mod
        
        # Clean sheet name for IDs
        sheet_id = sheet_name.replace(' ', '_').replace('(', '').replace(')', '')
        
        # Generate HTML for original side
        original_html = self._generate_table_html(
            original_df, column_headers, change_map, max_rows, max_cols, is_original=True
        )
        
        # Generate HTML for modified side
        modified_html = self._generate_table_html(
            modified_df, column_headers, change_map, max_rows, max_cols, is_original=False
        )
        
        # Create full HTML with embedded JavaScript for synchronized scrolling
        full_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{
                    margin: 0;
                    padding: 0;
                    font-family: 'Monaco', 'Menlo', 'Ubuntu Mono', monospace;
                    background: #1e1e1e;
                    color: #d4d4d4;
                }}
                
                .diff-container {{
                    background: #1e1e1e;
                    border-radius: 6px;
                    overflow: hidden;
                    height: 100vh;
                    display: flex;
                    flex-direction: column;
                }}
                
                .diff-header {{
                    background: #2d2d30;
                    color: #cccccc;
                    padding: 10px 15px;
                    font-weight: bold;
                    border-bottom: 1px solid #464647;
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    flex-shrink: 0;
                }}
                
                .sync-badge {{
                    background: #28a745;
                    color: white;
                    padding: 4px 10px;
                    border-radius: 4px;
                    font-size: 11px;
                    animation: pulse 2s infinite;
                }}
                
                @keyframes pulse {{
                    0% {{ opacity: 1; }}
                    50% {{ opacity: 0.7; }}
                    100% {{ opacity: 1; }}
                }}
                
                .diff-grid {{
                    display: grid;
                    grid-template-columns: 1fr 1fr;
                    gap: 2px;
                    background: #2d2d30;
                    padding: 1px;
                    flex: 1;
                    overflow: hidden;
                }}
                
                .diff-side {{
                    background: #1e1e1e;
                    display: flex;
                    flex-direction: column;
                    overflow: hidden;
                }}
                
                .diff-side-header {{
                    background: #2d2d30;
                    color: #969696;
                    padding: 8px 15px;
                    font-size: 12px;
                    border-bottom: 1px solid #464647;
                    flex-shrink: 0;
                }}
                
                .diff-content {{
                    overflow: auto;
                    flex: 1;
                }}
                
                .diff-table {{
                    display: table;
                    width: max-content;
                    min-width: 100%;
                    border-collapse: collapse;
                }}
                
                .diff-line {{
                    display: table-row;
                    min-height: 22px;
                    white-space: nowrap;
                }}
                
                .diff-line:hover {{
                    background: rgba(255, 255, 255, 0.05) !important;
                }}
                
                .line-number {{
                    display: table-cell;
                    background: #2d2d30;
                    color: #858585;
                    padding: 2px 8px;
                    min-width: 40px;
                    text-align: right;
                    border-right: 1px solid #464647;
                    position: sticky;
                    left: 0;
                    z-index: 5;
                    user-select: none;
                }}
                
                .line-content {{
                    display: table-cell;
                    padding: 2px 0;
                    white-space: nowrap;
                }}
                
                .cell-data {{
                    display: inline-block;
                    padding: 2px 12px;
                    min-width: 120px;
                    border-right: 1px solid #3a3a3a;
                    white-space: nowrap;
                }}
                
                .header-row {{
                    background: #2d2d30 !important;
                    font-weight: bold;
                    position: sticky;
                    top: 0;
                    z-index: 6;
                }}
                
                .header-row .cell-data {{
                    background: #2d2d30;
                    color: #007ACC;
                    font-weight: bold;
                    border-right: 1px solid #464647;
                }}
                
                .diff-modified {{
                    background: rgba(210, 153, 34, 0.15) !important;
                }}
                
                .diff-modified .line-number {{
                    background: #3d3319 !important;
                    color: #d29922;
                }}
                
                .cell-empty {{
                    color: #6a6a6a;
                    font-style: italic;
                    opacity: 0.6;
                }}
                
                .cell-value-old {{
                    background: rgba(229, 83, 75, 0.2);
                    color: #f85149;
                    padding: 1px 4px;
                    border-radius: 2px;
                    text-decoration: line-through;
                }}
                
                .cell-value-new {{
                    background: rgba(87, 171, 90, 0.2);
                    color: #57ab5a;
                    padding: 1px 4px;
                    border-radius: 2px;
                    font-weight: bold;
                }}
                
                /* Scrollbar styling */
                ::-webkit-scrollbar {{
                    width: 10px;
                    height: 10px;
                }}
                
                ::-webkit-scrollbar-track {{
                    background: #2d2d30;
                }}
                
                ::-webkit-scrollbar-thumb {{
                    background: #464647;
                    border-radius: 5px;
                }}
                
                ::-webkit-scrollbar-thumb:hover {{
                    background: #565658;
                }}
                
                ::-webkit-scrollbar-corner {{
                    background: #2d2d30;
                }}
            </style>
        </head>
        <body>
            <div class="diff-container">
                <div class="diff-header">
                    📊 Sheet: {sheet_name}
                    <span class="sync-badge">⟷ SYNCHRONIZED SCROLLING ACTIVE</span>
                </div>
                <div class="diff-grid">
                    <div class="diff-side">
                        <div class="diff-side-header">📁 Original</div>
                        <div class="diff-content" id="original-{sheet_id}">
                            {original_html}
                        </div>
                    </div>
                    <div class="diff-side">
                        <div class="diff-side-header">📝 Modified</div>
                        <div class="diff-content" id="modified-{sheet_id}">
                            {modified_html}
                        </div>
                    </div>
                </div>
            </div>
            
            <script>
                // Synchronized scrolling implementation
                (function() {{
                    const original = document.getElementById('original-{sheet_id}');
                    const modified = document.getElementById('modified-{sheet_id}');
                    
                    if (original && modified) {{
                        let isSyncing = false;
                        
                        // Function to sync scroll positions
                        function syncScroll(source, target) {{
                            if (!isSyncing) {{
                                isSyncing = true;
                                target.scrollTop = source.scrollTop;
                                target.scrollLeft = source.scrollLeft;
                                
                                // Reset flag after a small delay
                                requestAnimationFrame(() => {{
                                    isSyncing = false;
                                }});
                            }}
                        }}
                        
                        // Add scroll listeners
                        original.addEventListener('scroll', function() {{
                            syncScroll(original, modified);
                        }});
                        
                        modified.addEventListener('scroll', function() {{
                            syncScroll(modified, original);
                        }});
                        
                        // Log that synchronization is active
                        console.log('✅ Synchronized scrolling activated for sheet: {sheet_name}');
                    }} else {{
                        console.error('❌ Could not find elements for synchronization');
                    }}
                }})();
            </script>
        </body>
        </html>
        """
        
        return full_html
    
    def _generate_table_html(self, df, column_headers, change_map, max_rows, max_cols, is_original=True):
        """Generate HTML table for one side of the diff"""
        html_output = '<div class="diff-table">'
        
        # Add column headers
        html_output += '<div class="diff-line header-row"><span class="line-number">⬜</span><span class="line-content">'
        for col_idx in range(max_cols):
            col_num = col_idx + 1
            if col_num in column_headers:
                header = column_headers[col_num]
            elif col_idx < len(df.columns) if not df.empty else 0:
                header = df.columns[col_idx]
            else:
                header = get_column_letter(col_num)
            html_output += f'<span class="cell-data">{html.escape(str(header))}</span>'
        html_output += '</span></div>'
        
        # Add data rows
        for row_idx in range(max_rows):
            row_has_changes = any((row_idx, col) in change_map for col in range(max_cols))
            
            html_output += f'<div class="diff-line'
            if row_has_changes:
                html_output += ' diff-modified'
            html_output += f'"><span class="line-number">{row_idx + 1}</span><span class="line-content">'
            
            for col_idx in range(max_cols):
                cell_html = '<span class="cell-data">'
                
                if not df.empty and row_idx < len(df) and col_idx < len(df.columns):
                    value = df.iloc[row_idx, col_idx]
                    
                    if (row_idx, col_idx) in change_map:
                        if value is None:
                            cell_html += '<span class="cell-empty">[empty]</span>'
                        else:
                            if is_original:
                                cell_html += f'<span class="cell-value-old">{html.escape(str(value))}</span>'
                            else:
                                cell_html += f'<span class="cell-value-new">{html.escape(str(value))}</span>'
                    else:
                        if value is None:
                            cell_html += '<span class="cell-empty">·</span>'
                        else:
                            cell_html += html.escape(str(value))
                else:
                    cell_html += '<span class="cell-empty">·</span>'
                
                cell_html += '</span>'
                html_output += cell_html
            
            html_output += '</span></div>'
        
        html_output += '</div>'
        return html_output

# Main Streamlit App
def main():
    st.title("🔍 Excel Diff Visualizer")
    st.markdown("All changes shown as modifications - from original value to new value")
    
    # Sidebar for file upload
    with st.sidebar:
        st.header("📁 Upload Files")
        
        original_file = st.file_uploader(
            "Upload Original Template", 
            type=['xlsx', 'xls'],
            help="Upload the original Excel template"
        )
        
        modified_file = st.file_uploader(
            "Upload Modified Version", 
            type=['xlsx', 'xls'],
            help="Upload the modified Excel file"
        )
        
        st.divider()
        
        # View options
        st.header("⚙️ View Options")
        view_mode = st.radio(
            "Display Mode",
            ["Side-by-Side Diff", "Change List", "Summary Only"],
            index=0,
            help="Side-by-Side shows both versions with horizontal scrolling"
        )
        
        show_empty_changes = st.checkbox("Show [empty] → value changes", value=True)
        
        if st.button("🔍 Compare Files", type="primary", disabled=not (original_file and modified_file)):
            if original_file and modified_file:
                with st.spinner("Analyzing modifications..."):
                    st.session_state['original_file'] = original_file
                    st.session_state['modified_file'] = modified_file
                    
                    visualizer = ExcelDiffVisualizer(original_file, modified_file)
                    changes = visualizer.compare_sheets()
                    
                    st.session_state['changes'] = changes
                    st.session_state['visualizer'] = visualizer
                    st.session_state['comparison_done'] = True
    
    # Main content area
    if 'comparison_done' in st.session_state and st.session_state['comparison_done']:
        visualizer = st.session_state['visualizer']
        changes = st.session_state['changes']
        
        # Summary Statistics
        st.markdown("### 📊 Modification Summary")
        
        stats_html = f"""
        <div class="stats-container">
            <div class="stat-card">
                <div class="stat-number">{visualizer.summary['total_modifications']}</div>
                <div class="stat-label">Total Modifications</div>
            </div>
            <div class="stat-card from-empty">
                <div class="stat-number">{visualizer.summary['blank_to_value']}</div>
                <div class="stat-label">[empty] → Value</div>
            </div>
            <div class="stat-card to-empty">
                <div class="stat-number">{visualizer.summary['value_to_blank']}</div>
                <div class="stat-label">Value → [empty]</div>
            </div>
            <div class="stat-card modified">
                <div class="stat-number">{visualizer.summary['value_to_value']}</div>
                <div class="stat-label">Value → Value</div>
            </div>
        </div>
        """
        st.markdown(stats_html, unsafe_allow_html=True)
        
        # Sheet navigation
        if visualizer.summary['sheets_modified']:
            st.markdown("### 📑 Sheet Modifications")
            
            if view_mode == "Side-by-Side Diff":
                # VS Code style diff view with synchronized scrolling
                st.success("✅ **Synchronized Scrolling Active**: Both panels scroll together horizontally and vertically!")
                
                sheet_tabs = st.tabs(visualizer.summary['sheets_modified'])
                
                for idx, sheet_name in enumerate(visualizer.summary['sheets_modified']):
                    with sheet_tabs[idx]:
                        # Clean sheet name
                        clean_name = sheet_name.replace(" (New Sheet)", "").replace(" (Sheet Removed)", "")
                        
                        # Get the synchronized diff HTML
                        sheet_changes = changes.get(clean_name, {})
                        
                        if sheet_changes and not sheet_changes.get('sheet_added') and not sheet_changes.get('sheet_removed'):
                            # Use components for synchronized scrolling
                            diff_html = visualizer.create_synchronized_diff_component(clean_name)
                            components.html(diff_html, height=600, scrolling=False)
                        else:
                            # For added/removed sheets, use regular display
                            diff_html = visualizer.create_vs_code_diff_html(clean_name)
                            st.markdown(diff_html, unsafe_allow_html=True)
                        
                        # Quick stats for this sheet
                        if sheet_changes and sheet_changes.get('modifications'):
                            st.markdown("---")
                            
                            # Count change types
                            blank_to_value = sum(1 for m in sheet_changes['modifications'] if m['change_type'] == 'blank_to_value')
                            value_to_blank = sum(1 for m in sheet_changes['modifications'] if m['change_type'] == 'value_to_blank')
                            value_to_value = sum(1 for m in sheet_changes['modifications'] if m['change_type'] == 'value_to_value')
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("[empty] → Value", blank_to_value)
                            with col2:
                                st.metric("Value → [empty]", value_to_blank)
                            with col3:
                                st.metric("Value → Value", value_to_value)
            
            elif view_mode == "Change List":
                # List all modifications
                for sheet_name in visualizer.summary['sheets_modified']:
                    clean_name = sheet_name.replace(" (New Sheet)", "").replace(" (Sheet Removed)", "")
                    sheet_changes = changes.get(clean_name, {})
                    
                    with st.expander(f"📋 {sheet_name}", expanded=True):
                        if sheet_changes.get('modifications'):
                            # Group modifications by type
                            mods = sheet_changes['modifications']
                            
                            # Filter based on user preference
                            if not show_empty_changes:
                                mods = [m for m in mods if m['change_type'] == 'value_to_value']
                            
                            if mods:
                                st.markdown(f"**{len(mods)} modifications found:**")
                                
                                # Show first 20 modifications with column names
                                for mod in mods[:20]:
                                    old_val = visualizer._format_value(mod['old_value'])
                                    new_val = visualizer._format_value(mod['new_value'])
                                    
                                    # Use column name if available, otherwise use cell reference
                                    if 'column_name' in mod:
                                        location = f"{mod['column_name']}, Row {mod['row_number']}"
                                    else:
                                        location = f"Cell {mod['cell_ref']}"
                                    
                                    change_html = f"""
                                    <div class="change-item">
                                        <span class="change-location">{location}</span>
                                        <span class="change-arrow">→</span>
                                        Modified from <span class="value-old">{html.escape(old_val)}</span> 
                                        to <span class="value-new">{html.escape(new_val)}</span>
                                    </div>
                                    """
                                    st.markdown(change_html, unsafe_allow_html=True)
                                
                                if len(mods) > 20:
                                    st.info(f"... and {len(mods) - 20} more modifications")
                            else:
                                st.info("No modifications to display (check filter settings)")
            
            else:  # Summary Only
                st.markdown("### Summary View")
                summary_data = []
                
                for sheet_name in visualizer.summary['sheets_modified']:
                    clean_name = sheet_name.replace(" (New Sheet)", "").replace(" (Sheet Removed)", "")
                    sheet_changes = changes.get(clean_name, {})
                    
                    if sheet_changes.get('modifications'):
                        mods = sheet_changes['modifications']
                        
                        blank_to_value = sum(1 for m in mods if m['change_type'] == 'blank_to_value')
                        value_to_blank = sum(1 for m in mods if m['change_type'] == 'value_to_blank')
                        value_to_value = sum(1 for m in mods if m['change_type'] == 'value_to_value')
                        
                        summary_data.append({
                            'Sheet': sheet_name,
                            '[empty] → Value': blank_to_value,
                            'Value → [empty]': value_to_blank,
                            'Value → Value': value_to_value,
                            'Total': len(mods)
                        })
                
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    st.dataframe(
                        summary_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Sheet": st.column_config.TextColumn("Sheet Name", width="medium"),
                            "[empty] → Value": st.column_config.NumberColumn("Empty→Value", format="%d"),
                            "Value → [empty]": st.column_config.NumberColumn("Value→Empty", format="%d"),
                            "Value → Value": st.column_config.NumberColumn("Value→Value", format="%d"),
                            "Total": st.column_config.NumberColumn("Total Mods", format="%d"),
                        }
                    )
        else:
            st.info("✅ No modifications detected between the files")
        
        # Export section
        st.markdown("---")
        st.markdown("### 💾 Export Options")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("📄 Export Change Report", type="secondary"):
                # Generate detailed text report
                report = f"""EXCEL MODIFICATION REPORT
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SUMMARY
=======
Total Modifications: {visualizer.summary['total_modifications']}
- [empty] → Value: {visualizer.summary['blank_to_value']}
- Value → [empty]: {visualizer.summary['value_to_blank']}
- Value → Value: {visualizer.summary['value_to_value']}

DETAILED CHANGES BY SHEET
========================
"""
                for sheet_name in visualizer.summary['sheets_modified']:
                    clean_name = sheet_name.replace(" (New Sheet)", "").replace(" (Sheet Removed)", "")
                    sheet_changes = changes.get(clean_name, {})
                    
                    if sheet_changes.get('modifications'):
                        report += f"\n\nSheet: {sheet_name}\n"
                        report += "-" * (len(sheet_name) + 7) + "\n"
                        
                        for mod in sheet_changes['modifications']:
                            old_val = visualizer._format_value(mod['old_value'])
                            new_val = visualizer._format_value(mod['new_value'])
                            
                            # Use column name if available
                            if 'column_name' in mod:
                                location = f"{mod['column_name']}, Row {mod['row_number']}"
                            else:
                                location = mod['cell_ref']
                            
                            report += f"  {location}: {old_val} → {new_val}\n"
                
                st.download_button(
                    label="⬇️ Download Report",
                    data=report,
                    file_name=f"excel_modifications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                    mime="text/plain"
                )
        
        with col2:
            if st.button("📊 Export Diff Excel", type="secondary"):
                # Create diff Excel with highlighting
                diff_wb = openpyxl.Workbook()
                diff_wb.remove(diff_wb.active)
                
                # Color for modifications
                yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                
                for sheet_name, sheet_changes in changes.items():
                    if sheet_changes['sheet_removed']:
                        continue
                    
                    diff_sheet = diff_wb.create_sheet(sheet_name)
                    
                    # Use modified version as base
                    if sheet_name in visualizer.modified_wb.sheetnames:
                        source_sheet = visualizer.modified_wb[sheet_name]
                        
                        # Copy all cells and highlight modifications
                        for row in source_sheet.iter_rows():
                            for cell in row:
                                new_cell = diff_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                                
                                # Check if this cell was modified
                                for mod in sheet_changes.get('modifications', []):
                                    if mod['cell'] == (cell.row, cell.column):
                                        new_cell.fill = yellow_fill
                                        old_val = visualizer._format_value(mod['old_value'])
                                        new_val = visualizer._format_value(mod['new_value'])
                                        
                                        # Include column name in comment
                                        col_name = mod.get('column_name', mod['cell_ref'])
                                        comment_text = f"Location: {col_name}, Row {mod['row_number']}\n"
                                        comment_text += f"Modified from: {old_val}\n"
                                        comment_text += f"Modified to: {new_val}"
                                        
                                        new_cell.comment = openpyxl.comments.Comment(
                                            comment_text, 
                                            "Diff Tool"
                                        )
                                        break
                
                # Add summary sheet
                summary_sheet = diff_wb.create_sheet("MODIFICATION_SUMMARY", 0)
                summary_data = [
                    ["Excel Modification Report", ""],
                    ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    ["", ""],
                    ["Total Modifications:", visualizer.summary['total_modifications']],
                    ["[empty] → Value:", visualizer.summary['blank_to_value']],
                    ["Value → [empty]:", visualizer.summary['value_to_blank']],
                    ["Value → Value:", visualizer.summary['value_to_value']],
                    ["", ""],
                    ["Sheets Modified:", ", ".join(visualizer.summary['sheets_modified']) if visualizer.summary['sheets_modified'] else "None"],
                ]
                
                for row_idx, row_data in enumerate(summary_data, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                        if row_idx == 1:
                            cell.font = Font(bold=True, size=14)
                        elif col_idx == 1 and row_idx > 3:
                            cell.font = Font(bold=True)
                
                excel_buffer = io.BytesIO()
                diff_wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.download_button(
                    label="⬇️ Download Diff Excel",
                    data=excel_buffer,
                    file_name=f"excel_diff_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col3:
            if st.button("📋 Export JSON", type="secondary"):
                # Prepare JSON export
                change_report = {
                    'summary': visualizer.summary,
                    'modifications_by_sheet': {},
                    'timestamp': datetime.now().isoformat()
                }
                
                for sheet_name, sheet_changes in changes.items():
                    if sheet_changes.get('modifications'):
                        change_report['modifications_by_sheet'][sheet_name] = [
                            {
                                'cell': mod['cell_ref'],
                                'location': f"{mod.get('column_name', mod['cell_ref'])}, Row {mod['row_number']}",
                                'column': mod.get('column_name', get_column_letter(mod['col'])),
                                'row': mod['row_number'],
                                'from': visualizer._format_value(mod['old_value']),
                                'to': visualizer._format_value(mod['new_value']),
                                'type': mod['change_type']
                            }
                            for mod in sheet_changes['modifications']
                        ]
                
                json_str = json.dumps(change_report, indent=2, default=str)
                
                st.download_button(
                    label="⬇️ Download JSON",
                    data=json_str,
                    file_name=f"modifications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json"
                )
    
    else:
        # Welcome screen
        st.markdown("""
        ### Welcome to Excel Diff Visualizer! 
        
        This tool treats all Excel changes as **modifications** with enhanced features:
        
        #### 🎯 Key Features:
        - **Everything is a modification**: No "added" or "removed" - just "modified from X to Y"
        - **✨ SYNCHRONIZED SCROLLING**: Both panels scroll together (horizontal & vertical)
        - **Horizontal scrolling**: Clean table view with no line wrapping for wide spreadsheets
        - **Column header names**: Changes shown as "Q1 2024, Row 2" instead of "Cell B2"
        - **[empty] notation**: Clear indication when cells are blank
        
        #### 📊 Change Types:
        - 🔄 **[empty] → Value**: When a blank cell gets a value
        - 🔄 **Value → [empty]**: When a cell value is cleared
        - 🔄 **Value → Value**: When a cell value changes
        
        #### 🖥️ View Options:
        - **Side-by-Side Diff**: VS Code-style with **SYNCHRONIZED** scrolling (scroll one side, both move!)
        - **Change List**: Shows modifications with column names (e.g., "Product Name, Row 5")
        - **Summary View**: High-level statistics table
        
        **To get started:**
        1. Upload your original Excel template in the sidebar
        2. Upload the modified version from your client
        3. Click "Compare Files" to see all modifications
        
        The diff view now features **synchronized scrolling** - when you scroll one panel, the other 
        automatically follows, making it incredibly easy to compare changes across wide spreadsheets!
        """)

if __name__ == "__main__":
    main()
